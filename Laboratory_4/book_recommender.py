import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import json

tree = None  # Глобальная переменная для таблицы рекомендаций

# Загружает базу данных книг из JSON-файла
def load_books(filename="books.json"):
    with open(filename, "r", encoding="utf-8") as file:
        return json.load(file)

# Генерирует словарь предпочтений пользователя, включающий жанры, авторов и ключевые слова
def process_preferences(genres, authors, keywords):
    return {
        "genres": genres,
        "authors": authors,
        "keywords": keywords
    }

# Фильтрует книги по заданному жанру и году публикации
def filter_books(books, genre=None, year=None):
    return [
        book for book in books
        if (genre is None or book["genre"].lower() == genre.lower()) and
           (year is None or book["first_publish_year"] >= year)
    ]

# Рассчитывает "рейтинг соответствия" книги предпочтениям пользователя
def calculate_match(book, preferences):
    match_score = 0
    if book["genre"].lower() in (genre.lower() for genre in preferences["genres"]):
        match_score += 10
    if any(author.lower() in (auth.lower() for auth in preferences["authors"]) for author in book["author"]):
        match_score += 5
    if any(keyword.strip() and keyword.lower() in book["description"].lower() for keyword in preferences["keywords"]):
        match_score += 2
    return match_score

# Рекомендует книги на основе предпочтений пользователя и рассчитанных "рейтингом соответствия"
def recommend_books(books, preferences):
    rated_books = [(book, calculate_match(book, preferences)) for book in books]
    return sorted(rated_books, key=lambda x: x[1], reverse=True)

# Обновляет текстовую строку с выбранными авторами
def update_selected_authors(selected_authors, selected_authors_text):
    authors_text = ", ".join(selected_authors)
    selected_authors_text.set(authors_text)

# Обновляет список предложений авторов на основе введенного текста поиска
def update_author_suggestions(authors, author_search_entry, suggestions_frame, select_author):
    for widget in suggestions_frame.winfo_children():
        widget.destroy()

    query = author_search_entry.get().strip().lower()
    if not query:
        return

    matching_authors = [author for author in authors if query in author.lower()]
    for author in matching_authors[:20]:
        btn = tk.Button(
            suggestions_frame,
            text=author,
            command=lambda a=author: select_author(a),
            anchor="w",
            relief="flat",
            bg="#f0f0f0",
        )
        btn.config(width=70)
        btn.pack(fill="x", padx=5, pady=2)

# Обрабатывает выбор автора и добавляет/удаляет его из списка выбранных авторов
def select_author(author, selected_authors, selected_authors_text, author_search_entry, update_author_suggestions, authors, suggestions_frame):
    if author in selected_authors:
        response = messagebox.askyesno(
            "Подтверждение", f"Автор '{author}' уже выбран. Удалить его из списка?"
        )
        if response:
            selected_authors.remove(author)
    else:
        selected_authors.add(author)
    update_selected_authors(selected_authors, selected_authors_text)
    author_search_entry.delete(0, tk.END)
    update_author_suggestions(authors, author_search_entry, suggestions_frame, lambda a: select_author(a, selected_authors, selected_authors_text, author_search_entry, update_author_suggestions, authors, suggestions_frame))

# Отображает список рекомендаций в виде таблицы с книгами и их характеристиками
def show_recommendations(recommendations, results_frame):
    global tree

    for widget in results_frame.winfo_children():
        widget.destroy()

    columns = ("Название", "Автор(ы)", "Год", "Жанр", "Рейтинг соответствия")
    tree = ttk.Treeview(results_frame, columns=columns, show="headings", height=18)

    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, anchor="w", stretch=True, width=130)

    for book, score in recommendations:
        authors = ", ".join(book["author"])
        tree.insert("", "end", values=(book["title"], authors, book["first_publish_year"], book["genre"], score))

    scrollbar = ttk.Scrollbar(results_frame, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=scrollbar.set)

    tree.grid(row=0, column=0, sticky="nsew", pady=(6, 0))
    scrollbar.grid(row=0, column=1, sticky="ns", pady=(6, 0))

# Собирает и фильтрует данные, а затем запускает функцию для отображения рекомендаций
def get_recommendations(books, genre_vars, selected_authors, keywords_entry, year_from_entry, year_to_entry, sort_option, sort_order, only_selected_genres_var, show_recommendations, results_frame):
    selected_genres = [genre for genre, var in genre_vars.items() if var.get()]
    selected_authors_list = list(selected_authors)
    keywords = keywords_entry.get().split(", ")
    preferences = process_preferences(selected_genres, selected_authors_list, keywords)

    filtered_books = books if not only_selected_genres_var.get() else [book for book in books if book["genre"] in selected_genres]

    year_from = year_from_entry.get().strip()
    year_to = year_to_entry.get().strip()

    if year_from.isdigit() and year_to.isdigit():
        year_from = int(year_from)
        year_to = int(year_to)
        filtered_books = [book for book in filtered_books if year_from <= book.get("first_publish_year", 0) <= year_to]

    sort_reverse = (sort_order.get() == "desc")
    if sort_option.get() == "alphabet":
        filtered_books.sort(key=lambda x: x.get("title", "").lower(), reverse=sort_reverse)
    elif sort_option.get() == "year":
        filtered_books.sort(key=lambda x: x.get("first_publish_year", 0), reverse=sort_reverse)

    recommendations = recommend_books(filtered_books, preferences)
    show_recommendations(recommendations, results_frame)

# Сохраняет выбранные книги в Excel-файл
def save_to_read_list(tree, max_col_width=70):
    if tree is None:
        messagebox.showerror("Ошибка", "Книги не выбраны.")
        return
    
    selected_items = tree.selection()
    if not selected_items:
        messagebox.showwarning("Предупреждение", "Нет выбранных книг для сохранения")
        return

    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel файлы", "*.xlsx")])
    if file_path:
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Книги"

            headers = ["Название", "Автор(ы)", "Год", "Жанр", "Рейтинг соответствия"]
            sheet.append(headers)

            for item in selected_items:
                values = tree.item(item, "values")
                sheet.append(values)

            for col_idx, col in enumerate(sheet.columns, start=1):
                max_length = 0
                column = get_column_letter(col_idx)
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = min(max_length + 5, max_col_width)
                sheet.column_dimensions[column].width = adjusted_width

            workbook.save(file_path)
            messagebox.showinfo("Успех", "Данные успешно сохранены в Excel")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось сохранить файл: {e}")


root = tk.Tk()
root.title("Рекомендательная система книг")
root.geometry("1200x650")
root.resizable(False, False)

books = load_books()
genres = sorted({book["genre"] for book in books})
authors = sorted({author for book in books for author in book["author"]})

main_frame = tk.Frame(root)
main_frame.pack(padx=10, pady=5, fill="both", expand=True)

genres_frame = tk.LabelFrame(main_frame, text="Жанры", padx=10, pady=10)
genres_frame.grid(row=0, column=0, sticky="nsew")

genre_vars = {genre: tk.BooleanVar() for genre in genres}
num_columns = 4
for i, genre in enumerate(genres): tk.Checkbutton(genres_frame, text=genre, variable=genre_vars[genre]).grid(row=i // num_columns, column=i % num_columns, sticky="w", padx=5, pady=2)

only_selected_genres_var = tk.BooleanVar(value=False)
only_selected_genres_checkbutton = tk.Checkbutton(genres_frame, text="Рекомендовать только указанные жанры", variable=only_selected_genres_var)
only_selected_genres_checkbutton.grid(row=len(genres) // num_columns + 1, column=0, columnspan=num_columns, sticky="w", padx=5, pady=(0, 0))

years_frame = tk.LabelFrame(main_frame, text="Года", padx=5, pady=5)
years_frame.grid(row=1, column=0, sticky="nsew")

tk.Label(years_frame, text="Начиная с:").grid(row=0, column=0, padx=(10, 3), pady=(0, 4), sticky="w")
year_from_entry = tk.Entry(years_frame, width=10)
year_from_entry.grid(row=0, column=1, padx=0, pady=(0, 4))

tk.Label(years_frame, text="До:").grid(row=0, column=2, padx=(10, 3), pady=(0, 4), sticky="w")
year_to_entry = tk.Entry(years_frame, width=10)
year_to_entry.grid(row=0, column=3, padx=0, pady=(0, 4))

authors_frame = tk.LabelFrame(main_frame, text="Авторы", padx=0, pady=0)
authors_frame.grid(row=2, column=0, sticky="nsew")
authors_frame.grid_propagate(False)

author_selection_frame = tk.Frame(authors_frame)
author_selection_frame.pack(fill="x", pady=5)

tk.Label(author_selection_frame, text="Выбранные авторы:").pack(side="left", padx=(5, 10))
selected_authors_text = tk.StringVar(value="")
selected_authors_label = tk.Label(author_selection_frame, textvariable=selected_authors_text, wraplength=400, anchor="w", justify="left")
selected_authors_label.pack(side="left", fill="x", expand=True)

author_search_frame = tk.Frame(authors_frame)
author_search_frame.pack(fill="x", pady=5)

tk.Label(author_search_frame, text="Введите имя автора:").pack(side="left", padx=(5, 10))
author_search_entry = tk.Entry(author_search_frame)
author_search_entry.pack(side="left", fill="x", expand=True, padx=(5, 0))

canvas = tk.Canvas(authors_frame)
canvas.config(height=150)
scrollbar = tk.Scrollbar(authors_frame, orient="vertical", command=canvas.yview)
suggestions_frame = tk.Frame(canvas)
suggestions_frame.pack(fill="x", expand=True)

canvas.create_window((0, 0), window=suggestions_frame, anchor="nw")
canvas.config(yscrollcommand=scrollbar.set)

canvas.pack(side="left", fill="both", expand=True)
scrollbar.pack(side="right", fill="y")

suggestions_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

selected_authors = set()

author_search_entry.bind("<KeyRelease>", lambda e: update_author_suggestions(authors, author_search_entry, suggestions_frame, lambda a: select_author(a, selected_authors, selected_authors_text, author_search_entry, update_author_suggestions, authors, suggestions_frame)))

keywords_frame = tk.LabelFrame(main_frame, text="Ключевые слова", padx=10, pady=10)
keywords_frame.grid(row=3, column=0, columnspan=2, sticky="nsew")

keywords_entry = tk.Entry(keywords_frame)
keywords_entry.pack(fill="x")

results_frame = tk.LabelFrame(main_frame, text="Рекомендации", padx=0, pady=0)
results_frame.grid(row=0, column=1, rowspan=3, sticky="nsew")

results_canvas = tk.Canvas(results_frame)
results_inner_frame = tk.Frame(results_canvas)

results_canvas.create_window((0, 0), window=results_inner_frame, anchor="nw")

results_canvas.pack(fill="both", expand=True)

sort_frame = tk.LabelFrame(main_frame, text="Выбор сортировки", padx=10, pady=10)
sort_frame.grid(row=4, column=0, columnspan=3, sticky="nsew", pady=(10, 0))

sort_option = tk.StringVar(value="alphabet")
sort_order = tk.StringVar(value="asc")

tk.Radiobutton(sort_frame, text="По алфавиту", variable=sort_option, value="alphabet").grid(row=0, column=0, sticky="w", padx=5)
tk.Radiobutton(sort_frame, text="По году публикации", variable=sort_option, value="year").grid(row=0, column=1, sticky="w", padx=5)

tk.Radiobutton(sort_frame, text="Возрастание", variable=sort_order, value="asc").grid(row=1, column=0, sticky="w", padx=5)
tk.Radiobutton(sort_frame, text="Убывание", variable=sort_order, value="desc").grid(row=1, column=1, sticky="w", padx=5)

actions_frame = tk.Frame(main_frame)
actions_frame.grid(row=5, column=0, columnspan=2, sticky="nsew", padx=(0, 1))

main_frame.grid_rowconfigure(5, weight=1)
main_frame.grid_columnconfigure(1, weight=1)

actions_frame.grid_rowconfigure(0, weight=1)
actions_frame.grid_rowconfigure(1, weight=1)
actions_frame.grid_columnconfigure(0, weight=1)
actions_frame.grid_columnconfigure(1, weight=1)

save_button = tk.Button(actions_frame, text="Сохранить в xlsx", command=lambda: save_to_read_list(tree))
save_button.grid(row=0, column=0, columnspan=2, sticky="nsew")

recommend_button = tk.Button(
    actions_frame, 
    text="Получить рекомендации", 
    command=lambda: get_recommendations(
        books, genre_vars, selected_authors, keywords_entry, year_from_entry, 
        year_to_entry, sort_option, sort_order, only_selected_genres_var, 
        show_recommendations, results_inner_frame
    )
)
recommend_button.grid(row=1, column=0, columnspan=2, sticky="nsew")

root.mainloop() 